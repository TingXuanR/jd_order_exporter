#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, Iterable, List
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


RETURN_STATUS_COLUMN = "returnStatus"
RETURN_PRODUCT_INFO_COLUMN = "returnProductInfo"
ACTUAL_PAYMENT_AMOUNT_COLUMN = "actualPaymentAmount"
RETURN_PRODUCT_INFO_WIDTH = 24
RECEIVER_WIDTH = 8
REMOVED_COLUMNS = {"productImagePreview"}
INTEGER_COLUMNS = {"rangeLabel", "page", "productCount", "quantityTotal", RETURN_STATUS_COLUMN}
DECIMAL_COLUMNS = {"sequence", "amount", ACTUAL_PAYMENT_AMOUNT_COLUMN}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Merge jd-orders Excel files and enrich rows with return info "
            "from a JD after-sales CSV."
        )
    )
    parser.add_argument(
        "--orders-dir",
        default="jd-orders",
        help="Directory containing jd-orders Excel files and the after-sales CSV.",
    )
    parser.add_argument(
        "--afs-csv",
        default="jd-afs-20260513-025834.csv",
        help="After-sales CSV filename or path.",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Output .xlsx path. Default: <orders-dir>/jd-orders-merged-<timestamp>.xlsx",
    )
    return parser.parse_args()


def discover_order_files(orders_dir: Path, output_path: Path) -> List[Path]:
    files = []
    for path in sorted(orders_dir.glob("*.xlsx")):
        if path == output_path:
            continue
        if path.name.startswith("~$"):
            continue
        if path.name.startswith("jd-orders-merged-"):
            continue
        files.append(path)
    if not files:
        raise FileNotFoundError(f"No .xlsx files found under {orders_dir}")
    return files


def build_output_path(orders_dir: Path, output_arg: str) -> Path:
    if output_arg:
        return Path(output_arg).expanduser().resolve()
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return (orders_dir / f"jd-orders-merged-{stamp}.xlsx").resolve()


def read_afs_returns(csv_path: Path) -> Dict[str, Dict[str, object]]:
    by_order: Dict[str, Dict[str, object]] = {}
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if (row.get("售后类型") or "").strip() != "退货":
                continue
            if (row.get("列表状态") or "").strip() != "已完成":
                continue
            order_number = (row.get("订单号") or "").strip()
            if not order_number:
                continue
            product_name = (row.get("商品名称") or "").strip()
            refund_amount = parse_decimal(row.get("退款金额", "0"))

            entry = by_order.setdefault(
                order_number,
                {"refund_total": Decimal("0"), "products": []},
            )
            entry["refund_total"] = entry["refund_total"] + refund_amount
            if product_name and product_name not in entry["products"]:
                entry["products"].append(product_name)
    return by_order


def parse_decimal(value: object) -> Decimal:
    text = str(value or "").strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(quantized, ".2f")


def mask_receiver(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    name = text.split()[0]
    if len(name) <= 1:
        return "*"
    if len(name) == 2:
        return name[0] + "*"
    return name[0] + "**"


def convert_numeric_cell_value(header: str, value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return ""
    if header in INTEGER_COLUMNS:
        return int(parse_decimal(text))
    if header in DECIMAL_COLUMNS:
        return float(parse_decimal(text))
    return value


def combine_return_entries(*entries: Dict[str, object] | None) -> Dict[str, object] | None:
    refund_total = Decimal("0")
    products: List[str] = []
    found = False
    for entry in entries:
        if not entry:
            continue
        found = True
        refund_total += entry["refund_total"]
        for product in entry["products"]:
            if product and product not in products:
                products.append(product)
    if not found:
        return None
    return {"refund_total": refund_total, "products": products}


def build_parent_return_map(
    row_maps: List[Dict[str, object]],
    afs_returns: Dict[str, Dict[str, object]],
) -> Dict[str, Dict[str, object]]:
    parent_returns: Dict[str, Dict[str, object]] = {}
    for row_map in row_maps:
        sequence = str(row_map.get("sequence") or "").strip()
        if "." not in sequence:
            continue
        order_number = str(row_map.get("orderNumber") or "").strip()
        matched_return = afs_returns.get(order_number)
        if not matched_return:
            continue
        parent_sequence = sequence.split(".", 1)[0]
        parent_returns[parent_sequence] = combine_return_entries(
            parent_returns.get(parent_sequence),
            matched_return,
        )
    return parent_returns


def insert_return_columns(headers: List[str]) -> List[str]:
    output = [header for header in headers if header not in REMOVED_COLUMNS]
    if ACTUAL_PAYMENT_AMOUNT_COLUMN not in output:
        try:
            amount_index = output.index("amount")
        except ValueError as exc:
            raise KeyError("Expected an 'amount' column in the source Excel files") from exc
        output = output[: amount_index + 1] + [ACTUAL_PAYMENT_AMOUNT_COLUMN] + output[amount_index + 1 :]
    if RETURN_STATUS_COLUMN in output or RETURN_PRODUCT_INFO_COLUMN in output:
        return output
    try:
        status_index = output.index("status")
    except ValueError as exc:
        raise KeyError("Expected a 'status' column in the source Excel files") from exc
    return (
        output[: status_index + 1]
        + [RETURN_STATUS_COLUMN, RETURN_PRODUCT_INFO_COLUMN]
        + output[status_index + 1 :]
    )


def merge_rows(
    order_files: List[Path],
    afs_returns: Dict[str, Dict[str, object]],
) -> tuple[List[str], List[List[object]], Dict[int, float], Dict[str, object]]:
    merged_headers: List[str] | None = None
    merged_rows: List[List[object]] = []
    merged_row_heights: Dict[int, float] = {}
    layout_template: Dict[str, object] | None = None

    for path in order_files:
        try:
            workbook = load_workbook(path, data_only=True)
        except BadZipFile as exc:
            raise ValueError(f"{path.name} is not a valid .xlsx file") from exc
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            source_headers = [str(cell or "").strip() for cell in next(rows)]
            target_headers = insert_return_columns(source_headers)

            if merged_headers is None:
                merged_headers = target_headers
                layout_template = capture_layout_template(sheet, source_headers, target_headers)
            elif merged_headers != target_headers:
                raise ValueError(f"Header mismatch in {path.name}")

            file_row_maps: List[Dict[str, object]] = []
            file_row_numbers: List[int] = []
            for source_row_number, values in enumerate(rows, start=2):
                row_map = {}
                for index, header in enumerate(source_headers):
                    row_map[header] = values[index] if index < len(values) else None
                file_row_maps.append(row_map)
                file_row_numbers.append(source_row_number)

            parent_returns = build_parent_return_map(file_row_maps, afs_returns)
            for source_row_number, row_map in zip(file_row_numbers, file_row_maps):
                merged_rows.append(build_output_row(row_map, target_headers, afs_returns, parent_returns))
                row_height = sheet.row_dimensions[source_row_number].height
                if row_height is not None:
                    merged_row_heights[len(merged_rows) + 1] = row_height
        finally:
            workbook.close()

    if merged_headers is None or layout_template is None:
        raise ValueError("No rows were read from the Excel files")
    return merged_headers, merged_rows, merged_row_heights, layout_template


def build_output_row(
    row_map: Dict[str, object],
    headers: List[str],
    afs_returns: Dict[str, Dict[str, object]],
    parent_returns: Dict[str, Dict[str, object]],
) -> List[object]:
    sequence = str(row_map.get("sequence") or "").strip()
    order_number = str(row_map.get("orderNumber") or "").strip()
    amount_text = str(row_map.get("amount") or "").strip()
    amount_value = parse_decimal(amount_text)
    matched_return = afs_returns.get(order_number)
    parent_return = parent_returns.get(sequence) if sequence and "." not in sequence else None
    effective_return = combine_return_entries(matched_return, parent_return)
    display_return = matched_return

    return_status = 0
    return_product_info = ""
    actual_payment_amount = amount_text

    if effective_return:
        if amount_value != 0:
            actual_payment_amount = format_decimal(max(amount_value - effective_return["refund_total"], Decimal("0")))
    if display_return:
        return_status = 1
        return_product_info = " | ".join(display_return["products"])

    output = []
    for header in headers:
        if header == RETURN_STATUS_COLUMN:
            output.append(return_status)
        elif header == RETURN_PRODUCT_INFO_COLUMN:
            output.append(return_product_info)
        elif header == ACTUAL_PAYMENT_AMOUNT_COLUMN:
            output.append(convert_numeric_cell_value(header, actual_payment_amount))
        elif header == "receiver":
            output.append(mask_receiver(row_map.get(header, "")))
        else:
            output.append(convert_numeric_cell_value(header, row_map.get(header, "")))
    return output


def extract_cell_style(cell) -> Dict[str, object]:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def apply_cell_style(cell, style_spec: Dict[str, object] | None) -> None:
    if style_spec is None:
        return
    cell.font = copy(style_spec["font"])
    cell.fill = copy(style_spec["fill"])
    cell.border = copy(style_spec["border"])
    cell.alignment = copy(style_spec["alignment"])
    cell.number_format = style_spec["number_format"]
    cell.protection = copy(style_spec["protection"])


def capture_layout_template(
    sheet,
    source_headers: List[str],
    target_headers: List[str],
) -> Dict[str, object]:
    header_styles: Dict[str, object] = {}
    body_styles: Dict[str, object] = {}
    column_widths: Dict[str, float] = {}

    for column_index, header in enumerate(source_headers, start=1):
        column_letter = get_column_letter(column_index)
        column_width = sheet.column_dimensions[column_letter].width
        if column_width is not None:
            column_widths[header] = column_width
        header_styles[header] = extract_cell_style(sheet.cell(1, column_index))
        if sheet.max_row >= 2:
            body_styles[header] = extract_cell_style(sheet.cell(2, column_index))

    amount_width = column_widths.get("amount", 10)
    status_width = column_widths.get("status", 12)
    summary_width = column_widths.get("productSummary", 58)
    header_fallback = header_styles.get("status") or next(iter(header_styles.values()), None)
    amount_style = body_styles.get("amount")
    status_style = body_styles.get("status") or amount_style
    summary_style = body_styles.get("productSummary") or status_style
    receiver_style = body_styles.get("receiver") or status_style

    column_widths[ACTUAL_PAYMENT_AMOUNT_COLUMN] = amount_width
    column_widths[RETURN_STATUS_COLUMN] = status_width
    column_widths[RETURN_PRODUCT_INFO_COLUMN] = min(summary_width, RETURN_PRODUCT_INFO_WIDTH)
    column_widths["receiver"] = min(column_widths.get("receiver", RECEIVER_WIDTH), RECEIVER_WIDTH)

    if header_fallback is not None:
        header_styles[ACTUAL_PAYMENT_AMOUNT_COLUMN] = copy(header_fallback)
        header_styles[RETURN_STATUS_COLUMN] = copy(header_fallback)
        header_styles[RETURN_PRODUCT_INFO_COLUMN] = copy(header_fallback)
    if amount_style is not None:
        body_styles[ACTUAL_PAYMENT_AMOUNT_COLUMN] = copy(amount_style)
    if status_style is not None:
        body_styles[RETURN_STATUS_COLUMN] = copy(status_style)
    if summary_style is not None:
        body_styles[RETURN_PRODUCT_INFO_COLUMN] = copy(summary_style)
    if receiver_style is not None:
        body_styles["receiver"] = copy(receiver_style)

    default_data_height = None
    for row_index in range(2, min(sheet.max_row, 20) + 1):
        row_height = sheet.row_dimensions[row_index].height
        if row_height is not None:
            default_data_height = row_height
            break

    return {
        "headers": target_headers,
        "column_widths": column_widths,
        "header_styles": header_styles,
        "body_styles": body_styles,
        "header_height": sheet.row_dimensions[1].height,
        "default_data_height": default_data_height,
    }


def write_output(
    path: Path,
    headers: List[str],
    rows: List[List[object]],
    row_heights: Dict[int, float],
    layout_template: Dict[str, object],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "jd-orders-merged"
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index, header)
        apply_cell_style(cell, layout_template["header_styles"].get(header))

    for row_index, row in enumerate(rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, column_index, value)
            apply_cell_style(cell, layout_template["body_styles"].get(headers[column_index - 1]))

    for column_index, header in enumerate(headers, start=1):
        width = layout_template["column_widths"].get(header)
        if width is not None:
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    if layout_template["header_height"] is not None:
        sheet.row_dimensions[1].height = layout_template["header_height"]
    for row_index in range(2, len(rows) + 2):
        row_height = row_heights.get(row_index, layout_template["default_data_height"])
        if row_height is not None:
            sheet.row_dimensions[row_index].height = row_height
    workbook.save(path)


def main() -> None:
    args = parse_args()
    orders_dir = Path(args.orders_dir).expanduser().resolve()
    output_path = build_output_path(orders_dir, args.output)
    afs_csv_path = Path(args.afs_csv).expanduser()
    if not afs_csv_path.is_absolute():
        afs_csv_path = (orders_dir / afs_csv_path).resolve()

    order_files = discover_order_files(orders_dir, output_path)
    afs_returns = read_afs_returns(afs_csv_path)
    headers, rows, row_heights, layout_template = merge_rows(order_files, afs_returns)
    write_output(output_path, headers, rows, row_heights, layout_template)

    print(f"Merged {len(order_files)} files into: {output_path}")
    print(f"Rows written: {len(rows)}")
    print(f"Return orders matched: {len(afs_returns)}")
    print("Images copied: 0 (embedding disabled)")


if __name__ == "__main__":
    main()
