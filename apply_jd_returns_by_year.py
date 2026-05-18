#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import re
import xml.etree.ElementTree as ET
from copy import deepcopy
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


ACTUAL_PAYMENT_AMOUNT_COLUMN = "actualPaymentAmount"
RETURN_STATUS_COLUMN = "returnStatus"
RETURN_PRODUCT_INFO_COLUMN = "returnProductInfo"
RETURN_PRODUCT_INFO_WIDTH = 24.0
RECEIVER_WIDTH = 8.0
INTEGER_COLUMNS = {"rangeLabel", "page", "productCount", "quantityTotal", RETURN_STATUS_COLUMN}
DECIMAL_COLUMNS = {"sequence", "amount", ACTUAL_PAYMENT_AMOUNT_COLUMN}
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS = {"main": MAIN_NS}

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Filter completed JD return rows from the after-sales CSV and update "
            "yearly jd-orders workbooks without touching original image resources."
        )
    )
    parser.add_argument(
        "--orders-dir",
        default="jd-orders",
        help="Directory containing yearly jd-orders Excel files and the after-sales CSV.",
    )
    parser.add_argument(
        "--afs-csv",
        default="jd-afs-20260513-025834.csv",
        help="After-sales CSV filename or path.",
    )
    parser.add_argument(
        "--output-dir",
        default="",
        help="Directory for the filtered CSV and updated yearly workbooks.",
    )
    return parser.parse_args()


def parse_decimal(value: object) -> Decimal:
    text = str(value or "").strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(quantized, ".2f")


def combine_return_entries(*entries: Dict[str, object] | None) -> Dict[str, object] | None:
    refund_total = Decimal("0")
    products: List[str] = []
    found = False
    for entry in entries:
        if not entry:
            continue
        found = True
        refund_total += entry["refund_total"]
        for product in entry["products"]:
            if product and product not in products:
                products.append(product)
    if not found:
        return None
    return {"refund_total": refund_total, "products": products}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def build_output_dir(orders_dir: Path, output_arg: str) -> Path:
    if output_arg:
        return Path(output_arg).expanduser().resolve()
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return (orders_dir / f"jd-returns-by-year-{stamp}").resolve()


def read_completed_return_rows(csv_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if (row.get("售后类型") or "").strip() != "退货":
                continue
            if (row.get("列表状态") or "").strip() != "已完成":
                continue
            order_number = (row.get("订单号") or "").strip()
            if not order_number:
                continue
            rows.append({key: (value or "").strip() for key, value in row.items()})
    return rows


def aggregate_returns(rows: List[Dict[str, str]]) -> Dict[str, Dict[str, object]]:
    by_order: Dict[str, Dict[str, object]] = {}
    for row in rows:
        order_number = row["订单号"]
        product_name = row.get("商品名称", "")
        refund_amount = parse_decimal(row.get("退款金额", "0"))
        entry = by_order.setdefault(
            order_number,
            {"refund_total": Decimal("0"), "products": []},
        )
        entry["refund_total"] = entry["refund_total"] + refund_amount
        if product_name and product_name not in entry["products"]:
            entry["products"].append(product_name)
    return by_order


def write_filtered_csv(rows: List[Dict[str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "服务单号",
        "订单号",
        "申请时间",
        "商品名称",
        "售后类型",
        "列表状态",
        "退款金额",
        "结束时间",
        "来源列表页",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def discover_yearly_order_files(orders_dir: Path) -> List[Path]:
    files: List[Path] = []
    pattern = re.compile(r"^jd-orders-(20\d{2})-\d{8}-\d{6}\.xlsx$")
    for path in sorted(orders_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if pattern.match(path.name):
            files.append(path)
    if not files:
        raise FileNotFoundError(f"No yearly jd-orders .xlsx files found under {orders_dir}")
    return files


def extract_year_from_filename(path: Path) -> str:
    match = re.match(r"^jd-orders-(20\d{2})-\d{8}-\d{6}\.xlsx$", path.name)
    if not match:
        raise ValueError(f"Unsupported yearly workbook name: {path.name}")
    return match.group(1)


def cell_ref(column_index: int, row_number: int) -> str:
    return f"{get_column_letter(column_index)}{row_number}"


def split_cell_ref(ref: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    if not match:
        raise ValueError(f"Unexpected cell reference: {ref}")
    return column_index_from_string(match.group(1)), int(match.group(2))


def find_width_for_column(cols_elem, column_index: int) -> float:
    for col in cols_elem.findall(f"{{{MAIN_NS}}}col"):
        min_idx = int(col.attrib["min"])
        max_idx = int(col.attrib["max"])
        if min_idx <= column_index <= max_idx and "width" in col.attrib:
            return float(col.attrib["width"])
    return 13.0


def build_replacement_widths(cols_elem) -> List[float]:
    source_widths = [find_width_for_column(cols_elem, index) for index in range(1, 16)]
    return [
        *source_widths[:10],
        source_widths[9],
        source_widths[10],
        source_widths[11],
        source_widths[11],
        RETURN_PRODUCT_INFO_WIDTH,
        RECEIVER_WIDTH,
        source_widths[13],
        source_widths[14],
    ]


def mask_receiver(value: str) -> str:
    text = value.strip()
    if not text:
        return ""

    name = text.split()[0]
    if len(name) <= 1:
        return "*"
    if len(name) == 2:
        return name[0] + "*"
    return name[0] + "**"


def replace_cols(root) -> None:
    cols_elem = root.find("main:cols", NS)
    if cols_elem is None:
        raise ValueError("Worksheet is missing <cols>")

    new_widths = build_replacement_widths(cols_elem)
    new_cols = ET.Element(f"{{{MAIN_NS}}}cols")
    for index, width in enumerate(new_widths, start=1):
        col = ET.SubElement(new_cols, f"{{{MAIN_NS}}}col")
        col.set("min", str(index))
        col.set("max", str(index))
        col.set("width", format(width, "g"))
        col.set("customWidth", "1")

    root.remove(cols_elem)
    root.insert(0, new_cols)


def set_inline_text(cell_elem, text: str) -> None:
    cell_elem.set("t", "inlineStr")
    for child in list(cell_elem):
        cell_elem.remove(child)
    is_elem = ET.SubElement(cell_elem, f"{{{MAIN_NS}}}is")
    t_elem = ET.SubElement(is_elem, f"{{{MAIN_NS}}}t")
    t_elem.text = text


def set_numeric_text(cell_elem, text: str) -> None:
    cell_elem.attrib.pop("t", None)
    for child in list(cell_elem):
        cell_elem.remove(child)
    value_elem = ET.SubElement(cell_elem, f"{{{MAIN_NS}}}v")
    value_elem.text = text


def clone_cell(
    cell_elem,
    column_index: int,
    row_number: int,
    text: str | None = None,
    numeric: bool = False,
):
    new_cell = deepcopy(cell_elem)
    new_cell.set("r", cell_ref(column_index, row_number))
    if text is not None:
        if numeric:
            set_numeric_text(new_cell, text)
        else:
            set_inline_text(new_cell, text)
    return new_cell


def numeric_text_or_none(header: str, text: str) -> str | None:
    clean = text.strip()
    if not clean:
        return None
    if header in INTEGER_COLUMNS:
        return str(int(parse_decimal(clean)))
    if header in DECIMAL_COLUMNS:
        return str(parse_decimal(clean))
    return None


def extract_row_cells(row_elem) -> Dict[int, ET.Element]:
    cells: Dict[int, ET.Element] = {}
    for cell in row_elem.findall(f"{{{MAIN_NS}}}c"):
        col_index, _ = split_cell_ref(cell.attrib["r"])
        cells[col_index] = cell
    return cells


def build_parent_return_map(
    sheet_data,
    aggregated_returns: Dict[str, Dict[str, object]],
) -> Dict[str, Dict[str, object]]:
    parent_returns: Dict[str, Dict[str, object]] = {}
    for row_elem in sheet_data.findall(f"{{{MAIN_NS}}}row"):
        row_number = int(row_elem.attrib["r"])
        if row_number == 1:
            continue
        old_cells = extract_row_cells(row_elem)
        sequence = extract_inline_text(old_cells[3])
        if "." not in sequence:
            continue
        order_number = extract_inline_text(old_cells[5])
        matched_return = aggregated_returns.get(order_number)
        if not matched_return:
            continue
        parent_sequence = sequence.split(".", 1)[0]
        parent_returns[parent_sequence] = combine_return_entries(
            parent_returns.get(parent_sequence),
            matched_return,
        )
    return parent_returns


def rebuild_row(
    row_elem,
    aggregated_returns: Dict[str, Dict[str, object]],
    parent_returns: Dict[str, Dict[str, object]],
) -> bool:
    row_number = int(row_elem.attrib["r"])
    old_cells = extract_row_cells(row_elem)

    required_columns = list(range(1, 16))
    missing = [index for index in required_columns if index not in old_cells]
    if missing:
        raise ValueError(f"Row {row_number} is missing expected columns: {missing}")

    sequence = extract_inline_text(old_cells[3])
    order_number = extract_inline_text(old_cells[5])
    amount_text = extract_inline_text(old_cells[10])
    receiver_text = extract_inline_text(old_cells[13])
    amount_value = parse_decimal(amount_text)
    matched_return = aggregated_returns.get(order_number)
    parent_return = parent_returns.get(sequence) if sequence and "." not in sequence else None
    effective_return = combine_return_entries(matched_return, parent_return)
    display_return = matched_return

    actual_payment_amount = amount_text
    return_status = "0"
    return_product_info = ""
    if effective_return:
        if amount_value != 0:
            actual_payment_amount = format_decimal(max(amount_value - effective_return["refund_total"], Decimal("0")))
    if display_return:
        return_status = "1"
        return_product_info = " | ".join(display_return["products"])

    rebuilt_cells = [
        clone_cell(old_cells[1], 1, row_number, extract_inline_text(old_cells[1]), numeric=True),
        clone_cell(old_cells[2], 2, row_number, extract_inline_text(old_cells[2]), numeric=True),
        clone_cell(old_cells[3], 3, row_number, extract_inline_text(old_cells[3]), numeric=True),
        clone_cell(old_cells[4], 4, row_number),
        clone_cell(old_cells[5], 5, row_number),
        clone_cell(old_cells[6], 6, row_number),
        clone_cell(old_cells[7], 7, row_number),
        clone_cell(
            old_cells[8],
            8,
            row_number,
            extract_inline_text(old_cells[8]) if extract_inline_text(old_cells[8]).strip() else None,
            numeric=bool(extract_inline_text(old_cells[8]).strip()),
        ),
        clone_cell(
            old_cells[9],
            9,
            row_number,
            extract_inline_text(old_cells[9]) if extract_inline_text(old_cells[9]).strip() else None,
            numeric=bool(extract_inline_text(old_cells[9]).strip()),
        ),
        clone_cell(old_cells[10], 10, row_number, amount_text, numeric=True),
        clone_cell(old_cells[10], 11, row_number, actual_payment_amount, numeric=True),
        clone_cell(old_cells[11], 12, row_number),
        clone_cell(old_cells[12], 13, row_number),
        clone_cell(old_cells[12], 14, row_number, return_status, numeric=True),
        clone_cell(old_cells[7], 15, row_number, return_product_info),
        clone_cell(old_cells[13], 16, row_number, mask_receiver(receiver_text)),
        clone_cell(old_cells[14], 17, row_number),
        clone_cell(old_cells[15], 18, row_number),
    ]

    for child in list(row_elem):
        row_elem.remove(child)
    for cell in rebuilt_cells:
        row_elem.append(cell)
    row_elem.set("spans", "1:18")

    return effective_return is not None


def rebuild_header_row(row_elem) -> None:
    old_cells = {}
    for cell in row_elem.findall(f"{{{MAIN_NS}}}c"):
        col_index, _ = split_cell_ref(cell.attrib["r"])
        old_cells[col_index] = cell

    rebuilt_cells = [
        clone_cell(old_cells[1], 1, 1),
        clone_cell(old_cells[2], 2, 1),
        clone_cell(old_cells[3], 3, 1),
        clone_cell(old_cells[4], 4, 1),
        clone_cell(old_cells[5], 5, 1),
        clone_cell(old_cells[6], 6, 1),
        clone_cell(old_cells[7], 7, 1),
        clone_cell(old_cells[8], 8, 1),
        clone_cell(old_cells[9], 9, 1),
        clone_cell(old_cells[10], 10, 1),
        clone_cell(old_cells[10], 11, 1, ACTUAL_PAYMENT_AMOUNT_COLUMN),
        clone_cell(old_cells[11], 12, 1),
        clone_cell(old_cells[12], 13, 1),
        clone_cell(old_cells[12], 14, 1, RETURN_STATUS_COLUMN),
        clone_cell(old_cells[12], 15, 1, RETURN_PRODUCT_INFO_COLUMN),
        clone_cell(old_cells[13], 16, 1),
        clone_cell(old_cells[14], 17, 1),
        clone_cell(old_cells[15], 18, 1),
    ]

    for child in list(row_elem):
        row_elem.remove(child)
    for cell in rebuilt_cells:
        row_elem.append(cell)
    row_elem.set("spans", "1:18")


def extract_inline_text(cell_elem) -> str:
    inline_text = cell_elem.find(f"{{{MAIN_NS}}}is/{{{MAIN_NS}}}t")
    if inline_text is not None and inline_text.text is not None:
        return inline_text.text
    value_elem = cell_elem.find(f"{{{MAIN_NS}}}v")
    if value_elem is not None and value_elem.text is not None:
        return value_elem.text
    return ""


def patch_sheet_xml(sheet_xml: bytes, aggregated_returns: Dict[str, Dict[str, object]]) -> tuple[bytes, int]:
    root = ET.fromstring(sheet_xml)
    replace_cols(root)
    sheet_data = root.find("main:sheetData", NS)
    if sheet_data is None:
        raise ValueError("Worksheet is missing <sheetData>")

    parent_returns = build_parent_return_map(sheet_data, aggregated_returns)
    matched_rows = 0
    for row_elem in sheet_data.findall(f"{{{MAIN_NS}}}row"):
        row_number = int(row_elem.attrib["r"])
        if row_number == 1:
            rebuild_header_row(row_elem)
            continue
        if rebuild_row(row_elem, aggregated_returns, parent_returns):
            matched_rows += 1

    xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return xml_bytes, matched_rows


def update_yearly_workbook(
    workbook_path: Path,
    output_path: Path,
    aggregated_returns: Dict[str, Dict[str, object]],
) -> Dict[str, int]:
    try:
        with ZipFile(workbook_path, "r") as zin:
            sheet_xml = zin.read("xl/worksheets/sheet1.xml")
            patched_sheet_xml, matched_rows = patch_sheet_xml(sheet_xml, aggregated_returns)

            output_path.parent.mkdir(parents=True, exist_ok=True)
            with ZipFile(output_path, "w", compression=ZIP_DEFLATED) as zout:
                for info in zin.infolist():
                    data = patched_sheet_xml if info.filename == "xl/worksheets/sheet1.xml" else zin.read(info.filename)
                    zout.writestr(info, data)
    except BadZipFile as exc:
        raise ValueError(f"{workbook_path.name} is not a valid .xlsx file") from exc

    return {"matched_rows": matched_rows, "updated_rows": matched_rows}


def verify_workbook(output_path: Path) -> None:
    workbook = load_workbook(output_path, data_only=True)
    try:
        sheet = workbook.active
        headers = [normalize_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
        expected_headers = [
            "rangeLabel",
            "page",
            "sequence",
            "orderTime",
            "orderNumber",
            "productImagePreview",
            "productSummary",
            "productCount",
            "quantityTotal",
            "amount",
            ACTUAL_PAYMENT_AMOUNT_COLUMN,
            "paymentMethod",
            "status",
            RETURN_STATUS_COLUMN,
            RETURN_PRODUCT_INFO_COLUMN,
            "receiver",
            "detailUrl",
            "productImageUrls",
        ]
        if headers != expected_headers:
            raise ValueError(f"Unexpected headers in {output_path.name}: {headers}")
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()
    orders_dir = Path(args.orders_dir).expanduser().resolve()
    afs_csv_path = Path(args.afs_csv).expanduser()
    if not afs_csv_path.is_absolute():
        afs_csv_path = (orders_dir / afs_csv_path).resolve()
    output_dir = build_output_dir(orders_dir, args.output_dir)

    completed_return_rows = read_completed_return_rows(afs_csv_path)
    aggregated_returns = aggregate_returns(completed_return_rows)
    yearly_files = discover_yearly_order_files(orders_dir)

    output_dir.mkdir(parents=True, exist_ok=True)
    filtered_csv_path = output_dir / f"{afs_csv_path.stem}-completed-returns.csv"
    write_filtered_csv(completed_return_rows, filtered_csv_path)

    summary_lines = []
    for workbook_path in yearly_files:
        year = extract_year_from_filename(workbook_path)
        output_path = output_dir / f"{workbook_path.stem}-returns-updated.xlsx"
        stats = update_yearly_workbook(workbook_path, output_path, aggregated_returns)
        verify_workbook(output_path)
        summary_lines.append(
            f"{year}: {workbook_path.name} -> {output_path.name}, matched rows {stats['matched_rows']}"
        )

    print(f"Filtered completed returns CSV: {filtered_csv_path}")
    print(f"Completed return rows: {len(completed_return_rows)}")
    print(f"Unique returned orders: {len(aggregated_returns)}")
    for line in summary_lines:
        print(line)


if __name__ == "__main__":
    main()
